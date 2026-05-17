import json, csv, io
from datetime import datetime
from flask import (Blueprint, render_template, redirect, url_for,
                   session, jsonify, request, Response, send_file,
                   current_app, make_response)
from flask_login import login_required, current_user
from app.models import (Research, ResearchUser, Municipality, Property,
                         Contact, ActivityLog, User)
from app import db
from pathlib import Path
from app.utils import region_to_slug
from app.services.dashboard import build_dashboard_data

main_bp  = Blueprint('main', __name__)
DATA_DIR = Path(__file__).resolve().parents[1] / 'data'

def get_active_research():
    research_id = session.get('active_research_id')
    if research_id:
        r = Research.query.get(research_id)
        if r: return r
    m = ResearchUser.query.filter_by(user_id=current_user.id).first()
    if m:
        session['active_research_id'] = m.research_id
        return Research.query.get(m.research_id)
    return None

@main_bp.route('/')
@login_required
def index():
    research = get_active_research()
    if not research:
        return redirect(url_for('researches.list_researches'))

    muns = Municipality.query.filter_by(region=research.region).all()
    data = build_dashboard_data(research.id)

    # KPI specifici della dashboard che dipendono da muns (non inclusi nell'helper)
    data['kpi']['muns_explored'] = sum(1 for m in muns if m.research_status == 'explored')
    data['kpi']['total_muns']    = len(muns)

    logs = (ActivityLog.query.filter_by(research_id=research.id)
            .order_by(ActivityLog.timestamp.desc()).limit(20).all())
    user_ids  = {l.user_id for l in logs if l.user_id}
    users_map = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()}

    contacts = Contact.query.filter_by(research_id=research.id).all()

    return render_template('index.html',
        research=research, municipalities=muns,
        properties=data['props'], contacts=contacts,
        logs=logs, users_map=users_map,
        kpi=data['kpi'],
        chart_stati=data['chart_stati'],
        chart_prezzi=data['chart_prezzi'],
        chart_inserimenti=data['chart_inserimenti'])


@main_bp.route('/research/export-pdf')
@login_required
def export_research_pdf():
    import base64
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from weasyprint import HTML
    from io import BytesIO

    research = get_active_research()
    if not research:
        return redirect(url_for('researches.list_researches'))

    muns = Municipality.query.filter_by(region=research.region).all()
    data = build_dashboard_data(research.id)
    kpi  = data['kpi']
    kpi['total_muns']    = len(muns)
    kpi['muns_explored'] = sum(1 for m in muns if m.research_status == 'explored')

    def fig_to_b64(fig):
        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=120, bbox_inches='tight')
        buf.seek(0)
        plt.close(fig)
        return base64.b64encode(buf.read()).decode()

    # Grafico 1: donut stati
    chart_stati_img = None
    stati  = data['chart_stati']
    nonzero = [(l, v, c) for l, v, c in zip(stati['labels'], stati['data'], stati['colors']) if v > 0]
    if nonzero:
        fig, ax = plt.subplots(figsize=(4, 3.2))
        ax.pie([x[1] for x in nonzero], labels=[x[0] for x in nonzero],
               colors=[x[2] for x in nonzero], autopct='%1.0f%%',
               startangle=140, textprops={'fontsize': 8})
        chart_stati_img = fig_to_b64(fig)

    # Grafico 2: barre prezzo per comune
    chart_prezzi_img = None
    if data['chart_prezzi']['labels']:
        fig, ax = plt.subplots(figsize=(5, 3.2))
        bars = ax.barh(data['chart_prezzi']['labels'], data['chart_prezzi']['data'], color='#2D6A4F')
        ax.set_xlabel('€', fontsize=8)
        ax.tick_params(axis='both', labelsize=7)
        ax.invert_yaxis()
        chart_prezzi_img = fig_to_b64(fig)

    # Grafico 3: linea inserimenti
    chart_inserimenti_img = None
    if data['chart_inserimenti']['labels']:
        fig, ax = plt.subplots(figsize=(7, 2.8))
        ax.plot(data['chart_inserimenti']['labels'], data['chart_inserimenti']['data'],
                marker='o', color='#52B788', linewidth=2, markersize=5)
        ax.fill_between(data['chart_inserimenti']['labels'], data['chart_inserimenti']['data'],
                         alpha=0.15, color='#52B788')
        ax.tick_params(axis='x', rotation=45, labelsize=7)
        ax.tick_params(axis='y', labelsize=7)
        plt.tight_layout()
        chart_inserimenti_img = fig_to_b64(fig)

    props_sorted = sorted(data['props'], key=lambda p: p.created_at or datetime.min, reverse=True)[:50]

    html_str = render_template('report_research.html',
        research=research,
        author=current_user.full_name,
        now=datetime.utcnow().strftime('%d/%m/%Y %H:%M'),
        kpi=kpi,
        props=props_sorted,
        price_by_mun=data['price_by_mun'],
        chart_stati_img=chart_stati_img,
        chart_prezzi_img=chart_prezzi_img,
        chart_inserimenti_img=chart_inserimenti_img,
    )

    from flask import make_response
    pdf = HTML(string=html_str, base_url=current_app.root_path).write_pdf()
    response = make_response(pdf)
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=report-{research.id}.pdf'
    return response


@main_bp.route('/activity')
@login_required
def activity_log():
    research      = get_active_research()
    action_filter = request.args.get('action', '')
    q = ActivityLog.query.filter_by(research_id=research.id if research else -1)
    if action_filter: q = q.filter_by(action=action_filter)
    logs    = q.order_by(ActivityLog.timestamp.desc()).limit(200).all()
    actions = db.session.query(ActivityLog.action).filter_by(
        research_id=research.id if research else -1).distinct().all()
    user_ids  = {l.user_id for l in logs if l.user_id}
    users_map = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()}
    return render_template('activity_log.html', logs=logs,
                           action_filter=action_filter,
                           actions=[a[0] for a in actions],
                           users_map=users_map)


def _csv_safe(value):
    """Previene CSV injection prefissando con apostrofo celle che iniziano con =+-@."""
    s = str(value)
    return "'" + s if s and s[0] in ('=', '+', '-', '@') else s


@main_bp.route('/properties/export-csv')
@login_required
def export_properties_csv():
    from app.models import PropertyPhoto, PropertyContact as PC
    research = get_active_research()
    props    = Property.query.filter_by(research_id=research.id if research else -1).all()
    output   = io.StringIO()
    output.write('﻿')  # BOM UTF-8 per Excel italiano
    w = csv.writer(output)
    w.writerow(['ID', 'Titolo', 'Comune', 'Tipologia', 'Stato', 'Prezzo', 'Indirizzo',
                'URL annuncio', 'Foto', 'Contatti', 'Note', 'Data aggiunta'])
    for p in props:
        mun        = Municipality.query.get(p.municipality_id)
        photos     = PropertyPhoto.query.filter_by(property_id=p.id).all()
        photo_urls = ' | '.join(
            ph.external_url if ph.external_url else f'/uploads/{ph.file_path}' for ph in photos)
        links    = PC.query.filter_by(property_id=p.id).all()
        cont_ids = [l.contact_id for l in links]
        contacts = Contact.query.filter(Contact.id.in_(cont_ids)).all() if cont_ids else []
        w.writerow([
            p.id,
            _csv_safe(p.title or ''),
            mun.name if mun else '',
            p.property_type or '',
            p.status or '',
            p.price or '',
            _csv_safe(p.address or ''),
            p.listing_url or '',
            photo_urls,
            ' | '.join(c.name for c in contacts),
            _csv_safe((p.notes or '').replace('\n', ' ')),
            p.created_at.strftime('%d/%m/%Y %H:%M') if p.created_at else '',
        ])
    return Response(output.getvalue(), mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition':
            f'attachment; filename=immobili-{research.name if research else "fyp"}.csv'})


@main_bp.route('/properties/export-xlsx')
@login_required
def export_properties_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models import PropertyPhoto, PropertyContact as PC

    research = get_active_research()
    rid      = research.id if research else -1
    data     = build_dashboard_data(rid)
    props    = data['props']
    kpi      = data['kpi']

    wb = openpyxl.Workbook()

    # ── Foglio 1: Immobili ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Immobili'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2D6A4F')
    headers = ['ID', 'Titolo', 'Comune', 'Tipologia', 'Stato', 'Prezzo (€)',
               'Indirizzo', 'URL annuncio', 'Contatti', 'Note', 'Data aggiunta']
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = ws1.dimensions

    for p in props:
        mun      = Municipality.query.get(p.municipality_id)
        links    = PC.query.filter_by(property_id=p.id).all()
        cont_ids = [l.contact_id for l in links]
        contacts = Contact.query.filter(Contact.id.in_(cont_ids)).all() if cont_ids else []
        ws1.append([
            p.id,
            p.title or '',
            mun.name if mun else '',
            p.property_type or '',
            p.status or '',
            p.price or '',
            p.address or '',
            p.listing_url or '',
            ' | '.join(c.name for c in contacts),
            (p.notes or '').replace('\n', ' '),
            p.created_at.strftime('%d/%m/%Y %H:%M') if p.created_at else '',
        ])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = max(
            len(str(cell.value or '')) for cell in col) + 4

    # ── Foglio 2: KPI ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('KPI')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 18
    kpi_rows = [
        ('Totale immobili',       kpi['total']),
        ('Da valutare',           kpi['da_valutare']),
        ('Interessanti',          kpi['interessanti']),
        ('Offerta fatta',         kpi['offerta']),
        ('Scartati',              kpi['scartati']),
        ('Prezzo medio (€)',       kpi['avg_price'] or 'N/D'),
        ('Comuni con immobili',   kpi['comuni_coperti']),
    ]
    ws2.append(['Indicatore', 'Valore'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for row in kpi_rows:
        ws2.append(list(row))

    # ── Foglio 3: Per Comune ──────────────────────────────────────────────
    ws3 = wb.create_sheet('Per Comune')
    ws3.append(['Comune', 'Prezzo medio (€)'])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for name, avg in data['price_by_mun']:
        ws3.append([name, round(avg)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"immobili-{research.name if research else 'fyp'}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


@main_bp.route('/api/map-data')
@login_required
def map_data():
    research = get_active_research()
    if not research: return jsonify({'error': 'no_research'}), 400
    muns   = Municipality.query.filter_by(region=research.region).all()
    props  = Property.query.filter_by(research_id=research.id).all()
    conts  = Contact.query.filter_by(research_id=research.id).all()
    mun_status = {m.name: m.research_status for m in muns}
    slug      = region_to_slug(research.region)
    geo_path  = DATA_DIR / 'regions' / f'{slug}.geojson'
    if not geo_path.exists():
        geo_path = DATA_DIR / 'calabria_comuni.geojson'
    geojson = None
    if geo_path.exists():
        with open(geo_path, encoding='utf-8') as f:
            geojson = json.load(f)
        for feature in geojson['features']:
            name = feature['properties'].get('name', '')
            feature['properties']['status'] = mun_status.get(name, 'unexplored')
    return jsonify({
        'research':       {'id': research.id, 'name': research.name, 'region': research.region},
        'municipalities': [{'id': m.id, 'name': m.name, 'status': m.research_status,
                             'center_lat': m.center_lat, 'center_lng': m.center_lng} for m in muns],
        'properties':    [{'id': p.id, 'title': p.title, 'price': p.price,
                            'status': p.status, 'lat': p.lat, 'lng': p.lng,
                            'address': p.address, 'municipality_id': p.municipality_id} for p in props],
        'contacts':      [{'id': c.id, 'name': c.name, 'type': c.contact_type,
                            'phone': c.phone, 'lat': c.lat, 'lng': c.lng} for c in conts],
        'geojson':       geojson,
        'has_geojson':   geojson is not None
    })


@main_bp.route('/api/municipality/<int:mun_id>/status', methods=['POST'])
@login_required
def set_municipality_status(mun_id):
    if current_user.role not in ['admin', 'editor']:
        return jsonify({'error': 'forbidden'}), 403
    mun = Municipality.query.get_or_404(mun_id)
    new_status = request.json.get('status')
    if new_status in ['explored', 'partial', 'unexplored']:
        mun.research_status = new_status; db.session.commit()
        return jsonify({'ok': True, 'status': new_status})
    return jsonify({'error': 'invalid_status'}), 400


@main_bp.route('/switch-research/<int:research_id>')
@login_required
def switch_research(research_id):
    session['active_research_id'] = research_id
    return redirect(url_for('main.index'))

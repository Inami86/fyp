"""Test endpoint export. WeasyPrint richiede librerie di sistema su Linux;
i test PDF sono skippabili se l'ambiente non le ha."""
import io
import pytest

from app.models import Property


def test_csv_export_returns_csv(client, login_editor, sample_property):
    r = client.get('/properties/export-csv')
    assert r.status_code == 200
    assert r.mimetype == 'text/csv'
    body = r.data.decode('utf-8')
    # BOM UTF-8 deve essere il primo carattere
    assert body.startswith('﻿')
    # contiene la property creata
    assert 'Casa di Test' in body


def test_csv_export_injection_safe(client, login_editor, research, municipality, db):
    """Cella che inizia con = deve essere prefissata con apostrofo."""
    p = Property(research_id=research.id, municipality_id=municipality.id,
                 title='=cmd|calc', property_type='Altro', status='Da valutare')
    db.session.add(p)
    db.session.commit()
    r = client.get('/properties/export-csv')
    body = r.data.decode('utf-8')
    assert "'=cmd|calc" in body
    assert ',=cmd|calc,' not in body


def test_xlsx_export_returns_xlsx(client, login_editor, sample_property):
    r = client.get('/properties/export-xlsx')
    assert r.status_code == 200
    assert 'spreadsheetml.sheet' in r.mimetype
    # Magic bytes XLSX (zip)
    assert r.data[:2] == b'PK'


def test_xlsx_export_has_three_sheets(client, login_editor, sample_property):
    import openpyxl
    r = client.get('/properties/export-xlsx')
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert set(wb.sheetnames) == {'Immobili', 'KPI', 'Per Comune'}


def test_xlsx_kpi_matches_dashboard(client, login_editor, sample_property):
    import openpyxl
    r = client.get('/properties/export-xlsx')
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb['KPI']
    # Cerca "Totale immobili" e verifica il valore
    rows = list(ws.iter_rows(values_only=True))
    kpi_map = {row[0]: row[1] for row in rows if row[0]}
    assert kpi_map.get('Totale immobili') == 1
    assert kpi_map.get('Interessanti') == 1


def test_contacts_csv_export(client, login_editor, sample_contact):
    r = client.get('/contacts/export-csv')
    assert r.status_code == 200
    assert r.mimetype == 'text/csv'
    body = r.data.decode('utf-8')
    assert body.startswith('﻿')
    assert 'Mario Rossi' in body


def test_contacts_xlsx_export(client, login_editor, sample_contact):
    r = client.get('/contacts/export-xlsx')
    assert r.status_code == 200
    assert r.data[:2] == b'PK'


# ── PDF: skip se WeasyPrint non installato o librerie di sistema assenti ──

def _has_weasyprint():
    try:
        import weasyprint  # noqa: F401
        return True
    except (ImportError, OSError):
        return False


@pytest.mark.skipif(not _has_weasyprint(), reason='WeasyPrint non disponibile')
def test_property_pdf(client, login_editor, sample_property):
    r = client.get(f'/properties/{sample_property.id}/pdf')
    assert r.status_code == 200
    assert r.mimetype == 'application/pdf'
    assert r.data[:4] == b'%PDF'


@pytest.mark.skipif(not _has_weasyprint(), reason='WeasyPrint non disponibile')
def test_research_pdf(client, login_editor, sample_property):
    r = client.get('/research/export-pdf')
    assert r.status_code == 200
    assert r.mimetype == 'application/pdf'
    assert r.data[:4] == b'%PDF'
